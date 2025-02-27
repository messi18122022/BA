import os
import re
import statistics
import tkinter as tk
from tkinter import filedialog, messagebox

from PyPDF2 import PdfReader
from openpyxl import Workbook, load_workbook

def extract_data_from_pdf(pdf_path):
    """
    Liest den PDF-Inhalt ein und extrahiert alle Zeilen, die einen bekannten Analyt enthalten.
    Erwartete Analyten: Fluorid, Chlorid, Nitrit, Phosphat, Bromid, Nitrat, Sulfat.
    Für jede gefundene Zeile werden die Werte (t_R, A, N, Asym, w_base, R) extrahiert.
    Rückgabe: Liste von Listen, jeweils [Analyt, t_R, A, N, Asym, w_base, R]
    """
    analyte_set = {"fluorid", "chlorid", "nitrit", "phosphat", "bromid", "nitrat", "sulfat"}
    try:
        with open(pdf_path, "rb") as f:
            reader = PdfReader(f)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        messagebox.showerror("Fehler", f"Fehler beim Lesen der Datei {pdf_path}:\n{e}")
        return None

    rows = []
    for line in text.splitlines():
        tokens = line.strip().split()
        if tokens and tokens[0].lower() in analyte_set:
            if len(tokens) >= 7:
                try:
                    t_R    = float(tokens[1])
                    A      = float(tokens[2])
                    N      = float(tokens[3])
                    Asym   = float(tokens[4])
                    w_base = float(tokens[5])
                    R      = float(tokens[6])
                    rows.append([tokens[0], t_R, A, N, Asym, w_base, R])
                except ValueError:
                    continue
    return rows if rows else None

def extract_id_from_filename(filename):
    """
    Extrahiert die Messungs-ID aus dem Dateinamen.
    Erwartetes Format: YYYY-MM-DD_7STD_XX (Rest des Namens wird ignoriert)
    """
    basename = os.path.basename(filename)
    match = re.match(r"^(\d{4}-\d{2}-\d{2}_7STD_\d+)", basename)
    if match:
        return match.group(1)
    else:
        return os.path.splitext(basename)[0]

def main():
    # Tkinter initialisieren
    root = tk.Tk()
    root.withdraw()

    # Dialog zur Dateiauswahl
    file_paths = filedialog.askopenfilenames(
        title="Wählen Sie die PDF-Dateien aus",
        filetypes=[("PDF-Dateien", "*.pdf")]
    )
    if not file_paths:
        messagebox.showinfo("Abbruch", "Keine Dateien ausgewählt.")
        return

    # Hier werden alle Messdatensätze gespeichert:
    # Jeder Eintrag: (ID, Analyt, t_R, A, N, Asym, w_base, R)
    measurements = []

    for path in file_paths:
        data_rows = extract_data_from_pdf(path)
        if data_rows is None:
            messagebox.showerror("Fehler", f"In der Datei\n{path}\nkonnte keine Messreihe gefunden werden.")
            continue
        meas_id = extract_id_from_filename(path)
        for row in data_rows:
            # row: [Analyt, t_R, A, N, Asym, w_base, R]
            measurements.append((meas_id, row[0], row[1], row[2], row[3], row[4], row[5], row[6]))

    if not measurements:
        messagebox.showerror("Fehler", "Keine gültigen Messdaten gefunden.")
        return

    # Optional: Sortiere die Messungen nach der Injektionsnummer (angenommen, diese steht am Ende der ID)
    def get_injection_number(meas_id):
        match = re.search(r"_(\d+)$", meas_id)
        return int(match.group(1)) if match else 0

    measurements.sort(key=lambda x: get_injection_number(x[0]))

    # Gruppiere die Messungen für die Berechnung der Durchschnittswerte pro (Datum, Analyt)
    # Wir nehmen als gemeinsamen Präfix die ersten beiden Teile der ID (z. B. "2025-02-19_7STD")
    groups = {}
    for row in measurements:
        meas_id, analyt, t_R, A, N, Asym, w_base, R = row
        parts = meas_id.split("_")
        common_prefix = "_".join(parts[:2])
        key = (common_prefix, analyt)
        groups.setdefault(key, []).append((t_R, A, N, Asym, w_base, R))

    avg_rows = []
    for (common_prefix, analyt), values in groups.items():
        avg_std = []
        for i in range(6):
            col_values = [v[i] for v in values]
            mean_val = statistics.mean(col_values)
            std_val = statistics.stdev(col_values) if len(col_values) > 1 else 0
            avg_std.append(f"{mean_val:.3f} ± {std_val:.3f}")
        # Hier verwenden wir in der ID den gemeinsamen Präfix plus "_avg"
        avg_row = (f"{common_prefix}_avg", analyt, avg_std[0], avg_std[1], avg_std[2],
                   avg_std[3], avg_std[4], avg_std[5])
        avg_rows.append(avg_row)

    # Fester Pfad der Excel-Datei
    excel_path = '/Users/musamoin/Library/CloudStorage/OneDrive-ZHAW/01_Studium/Semester 6/BA/Messungen/Controlchart.xlsx'
    
    # Wenn die Datei existiert, laden und das Arbeitsblatt "Messungen" ersetzen, sonst neu erstellen
    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            if "Messungen" in wb.sheetnames:
                ws = wb["Messungen"]
                wb.remove(ws)
            ws = wb.create_sheet("Messungen")
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Laden der Excel-Datei:\n{e}")
            return
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Messungen"

    # Kopfzeile: ID, Analyt, t_R / min, A / ((µS/cm) * min), N, Asym, w_base, R
    headers = ["ID", "Analyt", "t_R / min", "A / ((µS/cm) * min)", "N", "Asym", "w_base", "R"]
    ws.append(headers)

    # Schreibe alle Messdatenzeilen
    for row in measurements:
        ws.append(row)

    # Füge die Durchschnittszeilen je Gruppe (pro (Datum, Analyt)) hinzu
    for row in avg_rows:
        ws.append(row)

    try:
        wb.save(excel_path)
        messagebox.showinfo("Erfolg", f"Die Excel-Datei wurde gespeichert:\n{excel_path}")
    except Exception as e:
        messagebox.showerror("Fehler", f"Fehler beim Speichern der Datei:\n{e}")

if __name__ == "__main__":
    main()
